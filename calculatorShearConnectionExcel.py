import openpyxl
import tkinter as tk
from tkinter import ttk

excel_file_path = "C:/Users/adria/Desktop/MyDB.xlsx"

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames
    return sheets

def read_sheet_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=2, values_only=True):
        if row[0] is not None:
            data.append(row[0])
    return data

def get_cell_value(file_path, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    selected_sheet = workbook[sheet_name]
    cell = selected_sheet.cell(row=row, column=column)
    return cell.value

def create_combobox(data):
    root = tk.Tk()
    root.title("Profiles")

    root.geometry("500x200+100+100")  # Ustaw rozmiar i położenie okna (szerokość x wysokość + pozycja x + pozycja y)
    root.minsize(500, 100)

    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)
    root.grid_rowconfigure(2, weight=1)
    root.grid_columnconfigure(0, weight=1)
    
    mainBeamFrame = ttk.Frame(root)
    mainBeamFrame.grid(row=0, column=0, sticky="nsew")
    
    mainBeamLbl = ttk.Label(mainBeamFrame, text="Main Beam", width=15)
    mainBeamLbl.grid(row=0, column=0, padx=2, pady=2)
    
    mainBeamProfileTypeCmbx = ttk.Combobox(mainBeamFrame, values=excel_sheets, state="readonly", width=5)
    mainBeamProfileTypeCmbx.grid(row=0, column=1, padx=2, pady=2)
    
    mainBeamProfileCmbx = ttk.Combobox(mainBeamFrame, values=data, state="readonly", width=10)
    mainBeamProfileCmbx.grid(row=0, column=2, padx=2, pady=2)

    mainBeamMaterialCmbx = ttk.Combobox(mainBeamFrame, values=["S235", "S355", "S275"], state="readonly", width=5)
    mainBeamMaterialCmbx.grid(row=0, column=3, padx=2, pady=2)
    mainBeamMaterialCmbx.current(0)
    
    secondaryBeamFrame = ttk.Frame(root)
    secondaryBeamFrame.grid(row=1, column=0, sticky="nsew")
    
    secondaryBeamLbl = ttk.Label(secondaryBeamFrame, text="Secondary Beam", width=15)
    secondaryBeamLbl.grid(row=0, column=0, padx=2, pady=2)
    
    secondaryBeamProfileTypeCmbx = ttk.Combobox(secondaryBeamFrame, values=excel_sheets, state="readonly", width=5)
    secondaryBeamProfileTypeCmbx.grid(row=0, column=1, padx=2, pady=2)
    
    secondaryBeamProfileCmbx = ttk.Combobox(secondaryBeamFrame, values=data, state="readonly", width=10)
    secondaryBeamProfileCmbx.grid(row=0, column=2, padx=2, pady=2)

    secondaryBeamMaterialCmbx = ttk.Combobox(secondaryBeamFrame, values=["S235", "S275", "S355"], state="readonly", width=5)
    secondaryBeamMaterialCmbx.grid(row=0, column=3, padx=2, pady=2)
    secondaryBeamMaterialCmbx.current(0)

    frame3 = ttk.Frame(root)
    frame3.grid(row=2, column=0, sticky="nsew")
    
    label3 = ttk.Label(frame3, text="x=")
    label3.grid(row=0, column=0, padx=2, pady=2)
    
    value_label = ttk.Label(frame3, text="")
    value_label.grid(row=0, column=1, padx=2, pady=2)
    
    def on_main_beam_profile_type_select(event):
        selected_sheet = mainBeamProfileTypeCmbx.get()
        selected_data = read_sheet_data(excel_file_path, selected_sheet)
        mainBeamProfileCmbx["values"] = selected_data
        mainBeamProfileCmbx.set("")
    
    mainBeamProfileTypeCmbx.bind("<<ComboboxSelected>>", on_main_beam_profile_type_select)
    
    def on_secondary_beam_profile_type_select(event):
        selected_sheet = secondaryBeamProfileTypeCmbx.get()
        selected_data = read_sheet_data(excel_file_path, selected_sheet)
        secondaryBeamProfileCmbx["values"] = selected_data
        secondaryBeamProfileCmbx.set("")
    
    secondaryBeamProfileTypeCmbx.bind("<<ComboboxSelected>>", on_secondary_beam_profile_type_select)
    
    def on_secondary_beam_profile_select(event):
        selected_sheet = secondaryBeamProfileTypeCmbx.get()
        selected_row_index = secondaryBeamProfileCmbx.current() + 2
        value = get_cell_value(excel_file_path, selected_sheet, selected_row_index, 1)
        value_label["text"] = value
    
    secondaryBeamProfileCmbx.bind("<<ComboboxSelected>>", on_secondary_beam_profile_select)

    root.mainloop()

excel_sheets = read_excel(excel_file_path)
create_combobox([])
